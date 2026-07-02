"""Excel 导出生成器"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def _style():
    """统一样式"""
    return {
        "title_font": Font(name="微软雅黑", size=14, bold=True),
        "header_font": Font(name="微软雅黑", size=10, bold=True),
        "header_fill": PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid"),
        "normal_font": Font(name="微软雅黑", size=10),
        "bold_font": Font(name="微软雅黑", size=10, bold=True),
        "amount_font": Font(name="微软雅黑", size=10, bold=True, color="1A73E8"),
        "thin_border": Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        ),
        "center_align": Alignment(horizontal="center", vertical="center"),
        "left_align": Alignment(horizontal="left", vertical="center"),
        "right_align": Alignment(horizontal="right", vertical="center"),
    }


def _apply_border(ws, min_row, max_row, min_col, max_col, center=True):
    s = _style()
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = s["thin_border"]
            if center:
                ws.cell(row=r, column=c).alignment = s["center_align"]


def _get_company_name():
    """从数据库获取公司名称（支持系统设置配置）"""
    try:
        from models import SystemConfig
        cfg = SystemConfig.query.filter_by(key="company_name").first()
        if cfg and cfg.value:
            return cfg.value
    except Exception:
        pass
    return "天津港航安装工程有限公司"


def new_report_wb(title, headers, col_widths=None):
    """创建汇总报表通用的 Workbook，返回 (wb, ws, s)"""
    s = _style()
    wb = Workbook()
    ws = wb.active

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].value = title
    ws["A1"].font = s["title_font"]
    ws["A1"].alignment = s["center_align"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = s["header_font"]
        cell.fill = s["header_fill"]
        cell.alignment = s["center_align"]
        cell.border = s["thin_border"]

    if col_widths:
        for col, w in col_widths:
            ws.column_dimensions[get_column_letter(col)].width = w

    return wb, ws, s

def generate_receiving_excel(note_data: dict, items_data: list, output_path: str):
    """生成材料点收单 Excel"""
    s = _style()
    wb = Workbook()
    ws = wb.active
    ws.title = "点收单"

    # 列宽
    for col, w in [(1, 8), (2, 28), (3, 12), (4, 8), (5, 12), (6, 16), (7, 16)]:
        ws.column_dimensions[get_column_letter(col)].width = w

    # 标题行
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = f"{_get_company_name()}《材料点收单》"
    c.font = s["title_font"]
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # 信息行（与打印页面一致）
    note_no = note_data.get("note_no", "")
    warehouse = note_data.get("warehouse_label", "")
    date = note_data.get("date", "")
    seller_name = note_data.get("seller_name", "")
    ws["A2"] = "点收编号:"
    ws["B2"] = note_no
    ws["D2"] = f"日期：{date}"
    ws["E2"] = f"仓库：{warehouse}"
    ws.merge_cells("D2:E2")
    ws["F2"] = "供应商："
    ws["G2"] = seller_name
    for cell in [ws["A2"], ws["B2"], ws["C2"], ws["D2"], ws["E2"], ws["F2"], ws["G2"]]:
        if cell.value:
            cell.font = s["normal_font"]

    # 表头
    headers = ["序号", "材料名称", "规格型号", "单位", "数量", "单价", "金额"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = s["header_font"]
        cell.fill = s["header_fill"]
    _apply_border(ws, 3, 3, 1, 7)

    # 数据行
    row = 4
    total_qty = 0
    total_amt = 0
    for i, item in enumerate(items_data, 1):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=item.get("material_name", ""))
        ws.cell(row=row, column=3, value=item.get("spec", ""))
        ws.cell(row=row, column=4, value=item.get("unit", ""))
        ws.cell(row=row, column=5, value=item.get("quantity", 0))
        ws.cell(row=row, column=5).number_format = '#,##0.000'
        ws.cell(row=row, column=6, value=item.get("unit_price", 0))
        ws.cell(row=row, column=6).number_format = '#,##0.00'
        ws.cell(row=row, column=7, value=item.get("amount", 0))
        ws.cell(row=row, column=7).number_format = '#,##0.00'
        total_qty += item.get("quantity", 0)
        total_amt += item.get("amount", 0)
        row += 1

    # 合计行（若没有明细也保留一行）
    if not items_data:
        row += 1

    ws.cell(row=row, column=1, value="合计")
    ws.cell(row=row, column=5, value=total_qty)
    ws.cell(row=row, column=5).number_format = '#,##0.000'
    ws.cell(row=row, column=7, value=total_amt)
    ws.cell(row=row, column=7).number_format = '#,##0.00'
    for c in range(1, 8):
        ws.cell(row=row, column=c).font = s["bold_font"]
    ws.cell(row=row, column=7).font = s["amount_font"]

    _apply_border(ws, 4, row, 1, 7, center=False)

    # 底部信息（与打印页面一致）
    row += 2
    ws.cell(row=row, column=1, value="工程编号").font = s["normal_font"]
    ws.merge_cells(f"B{row}:C{row}")
    ws.cell(row=row, column=2, value=note_data.get("project_no", ""))
    ws.cell(row=row, column=4, value="工程名称").font = s["normal_font"]
    ws.merge_cells(f"E{row}:G{row}")
    ws.cell(row=row, column=5, value=note_data.get("project_name", ""))
    row += 1

    # 记账人往右靠，与采购员间距缩小
    ws.cell(row=row, column=3, value="记账人").font = s["normal_font"]
    ws.cell(row=row, column=4, value=note_data.get("accountant", ""))
    ws.cell(row=row, column=5, value="采购员").font = s["normal_font"]
    ws.cell(row=row, column=6, value=note_data.get("buyer", ""))

    wb.save(output_path)
    return output_path


def generate_batch_workbook(title_prefix: str, notes, note_info_fn, output_path: str):
    """批量导出公共工厂函数

    Args:
        title_prefix: 标题前缀，如 "材料点收单"、"材料领用单"
        notes: 单据对象列表
        note_info_fn: callable(note) -> (note_data_dict, items_data_list)
        output_path: 输出路径
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = f"批量{title_prefix}"

    s_title = Font(name="微软雅黑", size=14, bold=True)
    s_header = Font(name="微软雅黑", size=10, bold=True)
    s_normal = Font(name="微软雅黑", size=10)
    fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    center = Alignment(horizontal="center", vertical="center")

    row = 1
    for note in notes:
        note_data, items_data = note_info_fn(note)

        # 标题
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        c = ws.cell(row=row, column=1, value=f"{title_prefix} - {note_data.get('note_no', '')}")
        c.font = s_title; c.alignment = center
        row += 1

        # 信息行
        info_text = note_data.get("info_line", "")
        if info_text:
            ws.cell(row=row, column=1, value=info_text).font = s_normal
            row += 1

        # 表头
        headers = ["序号", "材料名称", "规格型号", "单位", "数量", "单价", "金额"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = s_header; c.fill = fill; c.alignment = center; c.border = thin
        row += 1

        # 数据
        for i, item in enumerate(items_data, 1):
            ws.cell(row=row, column=1, value=i).font = s_normal
            ws.cell(row=row, column=1).alignment = center
            ws.cell(row=row, column=2, value=item.get("material_name", "")).font = s_normal
            ws.cell(row=row, column=3, value=item.get("spec", "")).font = s_normal
            ws.cell(row=row, column=4, value=item.get("unit", "")).font = s_normal
            ws.cell(row=row, column=4).alignment = center
            ws.cell(row=row, column=5, value=item.get("quantity", 0)).font = s_normal
            ws.cell(row=row, column=5).number_format = '#,##0.000'
            ws.cell(row=row, column=6, value=item.get("unit_price", 0)).font = s_normal
            ws.cell(row=row, column=6).number_format = '#,##0.00'
            ws.cell(row=row, column=7, value=item.get("amount", 0)).font = s_normal
            ws.cell(row=row, column=7).number_format = '#,##0.00'
            for c in range(1, 8):
                ws.cell(row=row, column=c).border = thin
            row += 1

        row += 2  # 每个单之间空行

    for col, w in [(1, 8), (2, 28), (3, 12), (4, 8), (5, 12), (6, 16), (7, 16)]:
        ws.column_dimensions[chr(64 + col)].width = w

    wb.save(output_path)
    return output_path


def generate_use_excel(note_data: dict, items_data: list, output_path: str):
    """生成材料领用单 Excel"""
    s = _style()
    wb = Workbook()
    ws = wb.active
    ws.title = "领用单"

    for col, w in [(1, 8), (2, 28), (3, 12), (4, 8), (5, 12), (6, 16), (7, 16)]:
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = f"{_get_company_name()}《材料领用单》"
    c.font = s["title_font"]
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    note_no = note_data.get("note_no", "")
    warehouse = note_data.get("warehouse_label", "")
    date = note_data.get("date", "")
    ws["A2"] = "领用编号:"
    ws["B2"] = note_no
    ws["D2"] = f"日期：{date}"
    ws.merge_cells("D2:E2")
    ws["F2"] = "仓库："
    ws["G2"] = warehouse
    for cell in [ws["A2"], ws["B2"], ws["C2"], ws["D2"], ws["E2"], ws["F2"], ws["G2"]]:
        if cell.value:
            cell.font = s["normal_font"]

    headers = ["序号", "材料名称", "规格型号", "单位", "数量", "单价", "金额"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = s["header_font"]
        cell.fill = s["header_fill"]
    _apply_border(ws, 3, 3, 1, 7)

    row = 4
    total_qty = 0
    total_amt = 0
    for i, item in enumerate(items_data, 1):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=item.get("material_name", ""))
        ws.cell(row=row, column=3, value=item.get("spec", ""))
        ws.cell(row=row, column=4, value=item.get("unit", ""))
        ws.cell(row=row, column=5, value=item.get("quantity", 0))
        ws.cell(row=row, column=5).number_format = '#,##0.000'
        ws.cell(row=row, column=6, value=item.get("unit_price", 0))
        ws.cell(row=row, column=6).number_format = '#,##0.00'
        ws.cell(row=row, column=7, value=item.get("amount", 0))
        ws.cell(row=row, column=7).number_format = '#,##0.00'
        total_qty += item.get("quantity", 0)
        total_amt += item.get("amount", 0)
        row += 1

    ws.cell(row=row, column=1, value="合计")
    ws.cell(row=row, column=5, value=total_qty)
    ws.cell(row=row, column=5).number_format = '#,##0.000'
    ws.cell(row=row, column=7, value=total_amt)
    ws.cell(row=row, column=7).number_format = '#,##0.00'
    for c in range(1, 8):
        ws.cell(row=row, column=c).font = s["bold_font"]
    ws.cell(row=row, column=7).font = s["amount_font"]
    _apply_border(ws, 4, row, 1, 7, center=False)

    row += 2
    ws.cell(row=row, column=1, value="工程编号").font = s["normal_font"]
    ws.merge_cells(f"B{row}:C{row}")
    ws.cell(row=row, column=2, value=note_data.get("project_no", ""))
    ws.cell(row=row, column=4, value="工程名称").font = s["normal_font"]
    ws.merge_cells(f"E{row}:G{row}")
    ws.cell(row=row, column=5, value=note_data.get("project_name", ""))
    row += 1
    ws.cell(row=row, column=3, value="记账人").font = s["normal_font"]
    ws.cell(row=row, column=4, value=note_data.get("accountant", ""))
    ws.cell(row=row, column=5, value="领用人").font = s["normal_font"]
    ws.cell(row=row, column=6, value=note_data.get("recipient", ""))

    wb.save(output_path)
    return output_path
