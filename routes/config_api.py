"""基础配置路由（仓库/工程/人员/报表/日志/设置）"""
from datetime import datetime
from flask import Blueprint, request, jsonify, session
from sqlalchemy import text
from models import Warehouse, Project, Personnel, Supplier, OperationLog, SystemConfig
from database import db
from services.operation_log import log_operation
from services.db_helpers import month_expr
from repositories.report_repository import (
    get_monthly_summary as _get_monthly_summary,
    get_project_summary as _get_project_summary,
    get_warehouse_summary as _get_warehouse_summary,
    get_supplier_summary as _get_supplier_summary,
    get_all_months as _get_all_months,
    get_warehouse_balance as _get_warehouse_balance,
    get_filter_options as _get_filter_options,
)
from repositories.inventory_repository import get_warehouse_name_map

bp = Blueprint("config", __name__, url_prefix="/api/v1")


# ===== 仓库管理 =====

@bp.route("/warehouses", methods=["GET"])
def list_warehouses():
    warehouses = Warehouse.query.order_by(Warehouse.code).all()
    return jsonify([{
        "id": w.id, "code": w.code, "name": w.name, "keywords": w.keywords
    } for w in warehouses])


@bp.route("/warehouses", methods=["POST"])
def create_warehouse():
    return jsonify({"error": "仓库分类为固定项，不允许新增"}), 403


@bp.route("/warehouses/<int:wh_id>", methods=["PUT"])
def update_warehouse(wh_id):
    if session.get("role") != "admin":
        return jsonify({"error": "仅管理员可执行此操作"}), 403
    w = Warehouse.query.get_or_404(wh_id)
    data = request.get_json()
    # 仓库名称/编码为固定项，只允许修改关键词
    if "name" in data and data["name"] != w.name:
        return jsonify({"error": "仓库名称为固定项，不允许修改"}), 403
    if "code" in data and data["code"] != w.code:
        return jsonify({"error": "仓库编码为固定项，不允许修改"}), 403
    w.keywords = data.get("keywords", w.keywords)
    db.session.commit()
    from services.warehouse_cache import invalidate_warehouse_cache
    invalidate_warehouse_cache()
    return jsonify({"message": "关键词已更新"})


@bp.route("/warehouses/<int:wh_id>", methods=["DELETE"])
def delete_warehouse(wh_id):
    return jsonify({"error": "仓库分类为固定项，不允许删除"}), 403


# ===== 工程管理 =====

def _project_sort_key(p):
    """按年份降序、分类升序、序号降序排序"""
    parts = p.project_no.split("-")
    if len(parts) == 3:
        return (-int(parts[2]), int(parts[0]), -int(parts[1]))
    return (0, 0, 0)

@bp.route("/projects/export", methods=["GET"])
def export_projects():
    """导出工程列表到 Excel"""
    import openpyxl
    from flask import send_file
    from io import BytesIO

    projects = sorted(Project.query.all(), key=_project_sort_key)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "工程列表"
    ws.append(["工程编号", "工程名称", "发包单位"])
    for p in projects:
        ws.append([p.project_no, p.project_name, p.client or ""])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="工程列表.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.route("/projects", methods=["GET"])
def list_projects():
    projects = sorted(Project.query.all(), key=_project_sort_key)
    return jsonify([{
        "id": p.id, "project_no": p.project_no, "project_name": p.project_name, "client": p.client or ""
    } for p in projects])


@bp.route("/projects/search", methods=["GET"])
def search_projects():
    q = request.args.get("q", "")
    if not q:
        return jsonify([])
    projects = Project.query.filter(
        db.or_(Project.project_no.contains(q), Project.project_name.contains(q))
    ).limit(10).all()
    return jsonify([{
        "id": p.id, "project_no": p.project_no, "project_name": p.project_name, "client": p.client or ""
    } for p in projects])


@bp.route("/projects", methods=["POST"])
def create_project():
    data = request.get_json()
    if Project.query.filter_by(project_no=data["project_no"]).first():
        return jsonify({"error": "工程编号已存在"}), 400
    p = Project(project_no=data["project_no"], project_name=data["project_name"], client=data.get("client", ""))
    db.session.add(p)
    db.session.commit()
    log_operation("create", "project", p.id, f"新增工程 {data['project_no']} - {data['project_name']}")
    return jsonify({"id": p.id, "message": "创建成功"})


@bp.route("/projects/import", methods=["POST"])
def import_projects():
    if session.get("role") != "admin":
        return jsonify({"error": "仅管理员可执行此操作"}), 403
    """从 Excel 导入工程列表"""
    if "file" not in request.files:
        return jsonify({"error": "请上传 Excel 文件"}), 400
    file = request.files["file"]
    if not file.filename.endswith((".xlsx", ".xls")):
        return jsonify({"error": "仅支持 .xlsx / .xls 格式"}), 400

    import openpyxl
    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    except Exception as e:
        return jsonify({"error": f"无法读取 Excel 文件: {e}"}), 400

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    # 跳过表头行，自动检测哪列是工程编号、哪列是工程名称
    header = next(rows_iter, None)
    if not header:
        return jsonify({"error": "Excel 文件为空"}), 400

    headers = [str(h).strip() if h else "" for h in header]
    idx_no = -1
    idx_name = -1
    idx_client = -1
    for i, h in enumerate(headers):
        if "编号" in h or "项目号" in h or "工程号" in h or "no" in h.lower() or "No" in h:
            idx_no = i
        if "名称" in h or "项目名" in h or "工程名" in h or "项目" in h:
            idx_name = i
        if "发包" in h or "业主" in h or "建设单位" in h or "客户" in h or "client" in h.lower():
            idx_client = i

    if idx_no < 0:
        idx_no = 0 if len(headers) > 0 else -1
    if idx_name < 0:
        idx_name = 1 if len(headers) > 1 else (0 if idx_no != 0 else -1)
    # 发包单位默认在第3列
    if idx_client < 0 and len(headers) > 2:
        idx_client = 2

    imported = 0
    skipped = 0
    errors = []

    for row_idx, row in enumerate(rows_iter, start=2):
        no = str(row[idx_no]).strip() if idx_no >= 0 and row[idx_no] is not None else ""
        name = str(row[idx_name]).strip() if idx_name >= 0 and row[idx_name] is not None else ""
        client = str(row[idx_client]).strip() if idx_client >= 0 and row[idx_client] is not None else ""
        if not no:
            continue
        if no.endswith(".0"):
            no = no[:-2]
        existing = Project.query.filter_by(project_no=no).first()
        if existing:
            if name:
                existing.project_name = name
            if client:
                existing.client = client
            skipped += 1
        else:
            p = Project(project_no=no, project_name=name or no, client=client)
            db.session.add(p)
        imported += 1

    db.session.commit()
    wb.close()
    return jsonify({
        "message": f"导入完成：新增 {imported - skipped} 个，跳过 {skipped} 个重复",
        "imported": imported,
        "skipped": skipped,
    })


@bp.route("/projects/<int:pid>", methods=["PUT"])
def update_project(pid):
    p = Project.query.get_or_404(pid)
    data = request.get_json()
    old_no = p.project_no
    old_name = p.project_name
    old_client = p.client or ""
    new_no = data.get("project_no", old_no)
    new_name = data.get("project_name", old_name)
    new_client = data.get("client", old_client)

    if new_no != old_no:
        conflict = Project.query.filter(Project.project_no == new_no, Project.id != pid).first()
        if conflict:
            return jsonify({"error": f"工程编号 '{new_no}' 已存在"}), 400

    p.project_no = new_no
    p.project_name = new_name
    p.client = new_client
    db.session.commit()

    # 同步到点收单和领用单
    if new_no != old_no or new_name != old_name or new_client != old_client:
        if new_no != old_no:
            db.session.execute(
                text("UPDATE receiving_notes SET project_no = :new WHERE project_no = :old"),
                {"new": new_no, "old": old_no}
            )
            db.session.execute(
                text("UPDATE use_notes SET project_no = :new WHERE project_no = :old"),
                {"new": new_no, "old": old_no}
            )
        if new_name != old_name:
            db.session.execute(
                text("UPDATE receiving_notes SET project_name = :new WHERE project_name = :old AND project_no = :no"),
                {"new": new_name, "old": old_name, "no": new_no}
            )
            db.session.execute(
                text("UPDATE use_notes SET project_name = :new WHERE project_name = :old AND project_no = :no"),
                {"new": new_name, "old": old_name, "no": new_no}
            )
        if new_client != old_client:
            db.session.execute(
                text("UPDATE receiving_notes SET client = :new WHERE project_no = :no"),
                {"new": new_client, "no": new_no}
            )
            db.session.execute(
                text("UPDATE use_notes SET client = :new WHERE project_no = :no"),
                {"new": new_client, "no": new_no}
            )
        db.session.commit()

    return jsonify({"message": "更新成功，已同步到相关单据"})


@bp.route("/projects/<int:pid>", methods=["DELETE"])
def delete_project(pid):
    if session.get("role") != "admin":
        return jsonify({"error": "仅管理员可删除工程"}), 403
    p = Project.query.get_or_404(pid)
    # 检查是否被点收单引用
    used = db.session.execute(
        text("SELECT COUNT(*) FROM receiving_notes WHERE project_no = :no AND status='active'"),
        {"no": p.project_no}
    ).scalar()
    if used:
        return jsonify({"error": f"工程 '{p.project_no}' 已被 {used} 张点收单使用，无法删除"}), 400
    used2 = db.session.execute(
        text("SELECT COUNT(*) FROM use_notes WHERE project_no = :no AND status='active'"),
        {"no": p.project_no}
    ).scalar()
    if used2:
        return jsonify({"error": f"工程 '{p.project_no}' 已被 {used2} 张领用单使用，无法删除"}), 400
    db.session.delete(p)
    log_operation("delete", "project", pid, f"删除工程 {p.project_no}")
    db.session.commit()
    return jsonify({"message": "删除成功"})


# ===== 人员管理 =====

@bp.route("/personnel", methods=["GET"])
def list_personnel():
    role = request.args.get("role", "")
    query = Personnel.query
    if role:
        query = query.filter(Personnel.role.contains(role))
    personnel = query.order_by(Personnel.role, Personnel.name).all()
    return jsonify([{
        "id": p.id, "name": p.name, "role": p.role, "enabled": p.enabled
    } for p in personnel])


@bp.route("/personnel", methods=["POST"])
def create_personnel():
    data = request.get_json()
    p = Personnel(name=data["name"], role=data.get("role", ""), enabled=data.get("enabled", True))
    db.session.add(p)
    db.session.commit()
    log_operation("create", "personnel", p.id, f'新增人员 {data["name"]}')
    return jsonify({"id": p.id, "message": "创建成功"})


@bp.route("/personnel/<int:pid>", methods=["PUT"])
def update_personnel(pid):
    p = Personnel.query.get_or_404(pid)
    data = request.get_json()
    p.name = data.get("name", p.name)
    p.role = data.get("role", p.role)
    p.enabled = data.get("enabled", p.enabled)
    db.session.commit()
    return jsonify({"message": "更新成功"})


@bp.route("/personnel/<int:pid>", methods=["DELETE"])
def delete_personnel(pid):
    if session.get("role") != "admin":
        return jsonify({"error": "仅管理员可删除人员"}), 403
    p = Personnel.query.get_or_404(pid)
    # 检查是否被点收单引用（采购员/记账人）
    used_recv = db.session.execute(
        text("SELECT COUNT(*) FROM receiving_notes WHERE (buyer = :name OR accountant = :name) AND status='active'"),
        {"name": p.name}
    ).scalar()
    if used_recv:
        return jsonify({"error": f"人员 '{p.name}' 已被 {used_recv} 张点收单引用，无法删除"}), 400
    used_use = db.session.execute(
        text("SELECT COUNT(*) FROM use_notes WHERE (recipient = :name OR accountant = :name) AND status='active'"),
        {"name": p.name}
    ).scalar()
    if used_use:
        return jsonify({"error": f"人员 '{p.name}' 已被 {used_use} 张领用单引用，无法删除"}), 400
    db.session.delete(p)
    log_operation("delete", "personnel", pid, f"删除人员 {p.name}")
    db.session.commit()
    return jsonify({"message": "删除成功"})


@bp.route("/personnel/template", methods=["GET"])
def download_personnel_template():
    """下载人员导入模板"""
    import openpyxl
    from flask import send_file
    from io import BytesIO
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "人员导入模板"
    ws.append(["姓名", "角色"])
    ws.append(["张三", "记账员"])
    ws.append(["李四", "采购员,领用人"])
    ws.append(["王五", "领用人"])
    # 角色提示
    ws["C1"] = "角色说明"
    ws["C2"] = "记账员 / 采购员 / 领用人（可多选，逗号分隔）"
    ws["C3"] = "也可以直接写英文: accountant,buyer / recipient"
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="人员导入模板.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.route("/personnel/import", methods=["POST"])
def import_personnel():
    """从 Excel 导入人员列表"""
    if "file" not in request.files:
        return jsonify({"error": "请上传 Excel 文件"}), 400
    file = request.files["file"]
    if not file.filename.endswith((".xlsx", ".xls")):
        return jsonify({"error": "仅支持 .xlsx / .xls 格式"}), 400

    import openpyxl
    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
    except Exception:
        return jsonify({"error": "无法读取 Excel 文件，请检查格式"}), 400
    finally:
        wb.close()

    ROLE_MAP = {"记账员": "accountant", "采购员": "buyer", "领用人": "recipient",
                "accountant": "accountant", "buyer": "buyer", "recipient": "recipient"}

    added = 0
    errors = []
    for i, row in enumerate(rows, 2):
        name = (row[0] or "").strip()
        role_raw = (row[1] or "").strip()
        if not name:
            continue
        if not role_raw:
            errors.append(f"第 {i} 行「{name}」缺少角色")
            continue
        roles = []
        for r in role_raw.replace("，", ",").split(","):
            r = r.strip()
            mapped = ROLE_MAP.get(r)
            if mapped:
                roles.append(mapped)
            else:
                errors.append(f"第 {i} 行「{name}」角色「{r}」无效")
        if not roles:
            continue
        if Personnel.query.filter_by(name=name).first():
            errors.append(f"第 {i} 行「{name}」已存在，跳过")
            continue
        db.session.add(Personnel(name=name, role=",".join(roles)))
        added += 1

    db.session.commit()
    log_operation("import", "personnel", 0, f"Excel导入 {added} 条人员记录")
    msg = f"成功导入 {added} 人"
    if errors:
        msg += "；" + "; ".join(errors[:5])
        if len(errors) > 5:
            msg += f"…(共{len(errors)}条错误)"
    return jsonify({"message": msg})


# ===== 供应商管理 =====

@bp.route("/suppliers/sync", methods=["POST"])
def sync_suppliers():
    if session.get("role") != "admin":
        return jsonify({"error": "仅管理员可执行此操作"}), 403
    """从点收单同步供应商名称到供应商管理表"""
    rows = db.session.execute(text("SELECT DISTINCT seller_name FROM receiving_notes WHERE seller_name != ''")).fetchall()
    added = 0
    for row in rows:
        name = row[0]
        if not Supplier.query.filter_by(name=name).first():
            db.session.add(Supplier(name=name))
            added += 1
    # 同步信用代码（从发票表补充）
    db.session.execute(text(
        "UPDATE suppliers SET credit_code = (SELECT seller_tax_no FROM invoices WHERE invoices.seller_name = suppliers.name AND seller_tax_no != '' LIMIT 1) "
        "WHERE credit_code = '' OR credit_code IS NULL"
    ))
    db.session.commit()
    return jsonify({"message": f"同步完成，新增 {added} 个供应商", "added": added})

@bp.route("/suppliers", methods=["GET"])
def list_suppliers():
    suppliers = Supplier.query.order_by(Supplier.name).all()
    return jsonify([{
        "id": s.id, "name": s.name, "credit_code": s.credit_code or "",
        "contact": s.contact, "phone": s.phone, "remark": s.remark,
    } for s in suppliers])


@bp.route("/suppliers", methods=["POST"])
def create_supplier():
    if session.get("role") != "admin":
        return jsonify({"error": "仅管理员可执行此操作"}), 403
    data = request.get_json()
    if Supplier.query.filter_by(name=data["name"]).first():
        return jsonify({"error": "供应商已存在"}), 400
    s = Supplier(name=data["name"], credit_code=data.get("credit_code", ""),
                 contact=data.get("contact", ""), phone=data.get("phone", ""),
                 remark=data.get("remark", ""))
    db.session.add(s)
    db.session.commit()
    log_operation("create", "supplier", s.id, f'新增供应商 {data["name"]}')
    return jsonify({"id": s.id, "message": "创建成功"})


@bp.route("/suppliers/<int:sid>", methods=["PUT"])
def update_supplier(sid):
    if session.get("role") != "admin":
        return jsonify({"error": "仅管理员可执行此操作"}), 403
    s = Supplier.query.get_or_404(sid)
    data = request.get_json()
    s.name = data.get("name", s.name)
    s.credit_code = data.get("credit_code", s.credit_code)
    s.contact = data.get("contact", s.contact)
    s.phone = data.get("phone", s.phone)
    s.remark = data.get("remark", s.remark)
    db.session.commit()
    return jsonify({"message": "更新成功"})


@bp.route("/suppliers/<int:sid>", methods=["DELETE"])
def delete_supplier(sid):
    if session.get("role") != "admin":
        return jsonify({"error": "仅管理员可删除供应商"}), 403
    s = Supplier.query.get_or_404(sid)
    db.session.delete(s)
    log_operation("delete", "supplier", sid, f"删除供应商 {s.name}")
    db.session.commit()
    return jsonify({"message": "删除成功"})


# ===== 编号计数器 =====

@bp.route("/counter/next", methods=["GET"])
def next_counter():
    from services.counter import peek_next_note_no
    r_no = peek_next_note_no("D")
    u_no = peek_next_note_no("L")
    return jsonify({"recv_note_no": r_no, "use_note_no": u_no})


# ===== 月度汇总报表 =====

@bp.route("/reports/monthly", methods=["GET"])
def monthly_report():
    """月度汇总报表（支持 months 逗号分隔筛选）"""
    months_param = request.args.get("months", "").strip()
    month_list = [m.strip() for m in months_param.split(",") if m.strip()] if months_param else []

    recv_rows, use_rows = _get_monthly_summary(db.session, month_list)
    proj_rows = _get_project_summary(db.session, month_list)
    wh_rows = _get_warehouse_summary(db.session, month_list)
    sup_rows = _get_supplier_summary(db.session, month_list)
    all_months = _get_all_months(db.session)

    return jsonify({
        "monthly_recv": [{"ym": r.ym, "cnt": r.cnt, "qty": round(float(r.qty), 3), "amt": round(float(r.amt), 2)} for r in recv_rows],
        "monthly_use": [{"ym": r.ym, "cnt": r.cnt, "qty": round(float(r.qty), 3), "amt": round(float(r.amt), 2)} for r in use_rows],
        "by_project": [{"name": r.project_name, "cnt": r.cnt, "amt": round(float(r.amt), 2)} for r in proj_rows],
        "by_warehouse": [{"code": r.warehouse_code, "name": r.wh_name or r.warehouse_code, "cnt": r.cnt, "amt": round(float(r.amt), 2)} for r in wh_rows],
        "by_supplier": [{"name": r.seller_name, "cnt": r.cnt, "amt": round(float(r.amt), 2)} for r in sup_rows],
        "all_months": all_months,
    })


@bp.route("/reports/monthly/export", methods=["GET"])
def export_monthly_report():
    """导出月度汇总报表 Excel"""
    import os
    from flask import current_app
    from services.excel_generator import new_report_wb

    recv_data, _ = _get_monthly_summary(db.session)

    wb, ws, s = new_report_wb("物资点收月度汇总报表", ["月份", "点收单数", "点收金额"], [(1, 15), (2, 12), (3, 16)])
    ws.title = "月度汇总"

    for i, r in enumerate(recv_data, 3):
        ws.cell(row=i, column=1, value=r.ym)
        ws.cell(row=i, column=2, value=r.cnt)
        c = ws.cell(row=i, column=3, value=round(float(r.amt), 2))
        c.number_format = '#,##0.00'
        for col in range(1, 4):
            ws.cell(row=i, column=col).font = s["normal_font"]
            ws.cell(row=i, column=col).border = s["thin_border"]
            if col == 1:
                ws.cell(row=i, column=col).alignment = s["center_align"]
            if col == 2:
                ws.cell(row=i, column=col).alignment = s["center_align"]

    export_dir = os.path.join(current_app.config.get("UPLOAD_FOLDER", "uploads"), "exports")
    os.makedirs(export_dir, exist_ok=True)
    output_path = os.path.join(export_dir, f"月度汇总报表_{datetime.now().strftime('%Y%m')}.xlsx")
    wb.save(output_path)

    return jsonify({"path": output_path, "filename": os.path.basename(output_path)})


@bp.route("/reports/by-project/export", methods=["GET"])
def export_by_project_report():
    """导出按工程汇总报表 Excel"""
    import os
    from flask import current_app
    from services.excel_generator import new_report_wb

    rows = _get_project_summary(db.session)

    wb, ws, s = new_report_wb("物资点收按工程汇总报表", ["工程名称", "点收单数", "点收金额"], [(1, 30), (2, 12), (3, 16)])
    ws.title = "按工程汇总"

    for i, r in enumerate(rows, 3):
        ws.cell(row=i, column=1, value=r.project_name).font = s["normal_font"]
        ws.cell(row=i, column=2, value=r.cnt).font = s["normal_font"]
        ws.cell(row=i, column=2).alignment = s["center_align"]
        c = ws.cell(row=i, column=3, value=round(float(r.amt), 2))
        c.font = s["normal_font"]
        c.number_format = '#,##0.00'
        for col in range(1, 4):
            ws.cell(row=i, column=col).border = s["thin_border"]

    export_dir = os.path.join(current_app.config.get("UPLOAD_FOLDER", "uploads"), "exports")
    os.makedirs(export_dir, exist_ok=True)
    output_path = os.path.join(export_dir, f"按工程汇总报表_{datetime.now().strftime('%Y%m')}.xlsx")
    wb.save(output_path)

    return jsonify({"path": output_path, "filename": os.path.basename(output_path)})


@bp.route("/reports/by-warehouse/export", methods=["GET"])
def export_by_warehouse_report():
    """导出按仓库汇总报表 Excel"""
    import os
    from flask import current_app
    from services.excel_generator import new_report_wb

    rows = _get_warehouse_summary(db.session)

    wb, ws, s = new_report_wb("物资点收按仓库汇总报表", ["仓库编码", "仓库名称", "点收单数", "点收金额"], [(1, 10), (2, 20), (3, 12), (4, 16)])
    ws.title = "按仓库汇总"

    for i, r in enumerate(rows, 3):
        ws.cell(row=i, column=1, value=r.warehouse_code)
        ws.cell(row=i, column=2, value=r.wh_name or r.warehouse_code)
        ws.cell(row=i, column=3, value=r.cnt)
        c = ws.cell(row=i, column=4, value=round(float(r.amt), 2))
        c.number_format = '#,##0.00'
        for col in range(1, 5):
            ws.cell(row=i, column=col).font = s["normal_font"]
            ws.cell(row=i, column=col).border = s["thin_border"]
            if col <= 3:
                ws.cell(row=i, column=col).alignment = s["center_align"]

    export_dir = os.path.join(current_app.config.get("UPLOAD_FOLDER", "uploads"), "exports")
    os.makedirs(export_dir, exist_ok=True)
    output_path = os.path.join(export_dir, f"按仓库汇总报表_{datetime.now().strftime('%Y%m')}.xlsx")
    wb.save(output_path)

    return jsonify({"path": output_path, "filename": os.path.basename(output_path)})


@bp.route("/reports/by-supplier/export", methods=["GET"])
def export_by_supplier_report():
    """导出按供应商汇总报表 Excel"""
    import os
    from flask import current_app
    from services.excel_generator import new_report_wb

    rows = _get_supplier_summary(db.session)

    wb, ws, s = new_report_wb("物资点收按供应商汇总报表", ["供应商名称", "点收单数", "点收金额"], [(1, 30), (2, 12), (3, 16)])
    ws.title = "按供应商汇总"

    for i, r in enumerate(rows, 3):
        ws.cell(row=i, column=1, value=r.seller_name)
        ws.cell(row=i, column=2, value=r.cnt)
        c = ws.cell(row=i, column=3, value=round(float(r.amt), 2))
        c.number_format = '#,##0.00'
        for col in range(1, 4):
            ws.cell(row=i, column=col).font = s["normal_font"]
            ws.cell(row=i, column=col).border = s["thin_border"]
            if col <= 2:
                ws.cell(row=i, column=col).alignment = s["center_align"]

    export_dir = os.path.join(current_app.config.get("UPLOAD_FOLDER", "uploads"), "exports")
    os.makedirs(export_dir, exist_ok=True)
    output_path = os.path.join(export_dir, f"按供应商汇总报表_{datetime.now().strftime('%Y%m')}.xlsx")
    wb.save(output_path)

    return jsonify({"path": output_path, "filename": os.path.basename(output_path)})


# ===== 统一汇总报表查询（方案A）=====

def _build_report_filters(args, prefix=""):
    """从请求参数构建筛选条件，prefix 为表别名前缀（如 'r.'）
    返回 (where_sql, params_dict)"""
    wheres = []
    params = {}
    p = prefix + "." if prefix else ""

    years = args.get("years", "").strip()
    months = args.get("months", "").strip()
    project_no = args.get("project_no", "").strip()
    warehouse_code = args.get("warehouse_code", "").strip()
    supplier_name = args.get("supplier_name", "").strip()

    if years:
        ys = [y.strip() for y in years.split(",") if y.strip()]
        if ys:
            clauses = " OR ".join([f"{p}date LIKE :y{i}" for i in range(len(ys))])
            wheres.append(f"({clauses})")
            for i, y in enumerate(ys):
                params[f"y{i}"] = f"{y}-%"

    if months:
        ms = [m.strip() for m in months.split(",") if m.strip()]
        if ms:
            clauses = " OR ".join([f"{month_expr(p + 'date')} LIKE :m{i}" for i in range(len(ms))])
            wheres.append(f"({clauses})")
            for i, m in enumerate(ms):
                params[f"m{i}"] = f"%{m}"

    if project_no:
        p_list = [x.strip() for x in project_no.split(",") if x.strip()]
        if len(p_list) == 1:
            wheres.append(f"{p}project_no = :project_no")
            params["project_no"] = p_list[0]
        elif len(p_list) > 1:
            clauses = " OR ".join([f"{p}project_no = :pn{i}" for i in range(len(p_list))])
            wheres.append(f"({clauses})")
            for i, pv in enumerate(p_list):
                params[f"pn{i}"] = pv

    if warehouse_code:
        w_list = [x.strip() for x in warehouse_code.split(",") if x.strip()]
        if len(w_list) == 1:
            wheres.append(f"{p}warehouse_code = :warehouse_code")
            params["warehouse_code"] = w_list[0]
        elif len(w_list) > 1:
            clauses = " OR ".join([f"{p}warehouse_code = :wc{i}" for i in range(len(w_list))])
            wheres.append(f"({clauses})")
            for i, wv in enumerate(w_list):
                params[f"wc{i}"] = wv

    if supplier_name:
        s_list = [x.strip() for x in supplier_name.split(",") if x.strip()]
        if len(s_list) == 1:
            wheres.append(f"{p}seller_name = :supplier_name")
            params["supplier_name"] = s_list[0]
        elif len(s_list) > 1:
            clauses = " OR ".join([f"{p}seller_name = :sn{i}" for i in range(len(s_list))])
            wheres.append(f"({clauses})")
            for i, sv in enumerate(s_list):
                params[f"sn{i}"] = sv

    where_sql = " AND ".join(wheres) if wheres else "1=1"
    return where_sql, params


@bp.route("/reports/query", methods=["GET"])
def report_query():
    """统一报表查询 — 支持按年/月/工程/仓库/供应商多维筛选"""
    recv_where, params = _build_report_filters(request.args, "r")
    # use_notes 没有 seller_name，去除该参数
    use_args = {k: v for k, v in request.args.items() if k != "supplier_name"}
    use_where, _ = _build_report_filters(use_args, "u")

    # 月份表达式（带别名）
    ym_expr_recv = f"{month_expr('r.date')} as ym"
    ym_expr_use = f"{month_expr('u.date')} as ym"

    # 点收汇总（含数量合计）
    recv_sql = text(f"""
        SELECT {ym_expr_recv}, r.project_no, r.project_name,
               r.warehouse_code, r.seller_name,
               COUNT(*) as cnt, SUM(r.total_amount) as amt,
               COALESCE(SUM(riq.note_qty), 0) as qty
        FROM receiving_notes r
        LEFT JOIN (SELECT note_id, SUM(quantity) as note_qty FROM receiving_items GROUP BY note_id) riq ON riq.note_id = r.id
        WHERE r.status='active' AND {recv_where}
        GROUP BY {month_expr('r.date')}, r.project_no, r.project_name, r.warehouse_code, r.seller_name
        ORDER BY {month_expr('r.date')} DESC, r.project_no
    """)
    recv_rows = db.session.execute(recv_sql, params).fetchall()

    # 领用汇总（含数量合计）
    use_sql = text(f"""
        SELECT {ym_expr_use}, u.project_no, u.project_name,
               u.warehouse_code, COALESCE(r.seller_name,'') as seller_name,
               COUNT(*) as cnt, SUM(u.total_amount) as amt,
               COALESCE(SUM(uiq.note_qty), 0) as qty
        FROM use_notes u
        LEFT JOIN (SELECT note_id, SUM(quantity) as note_qty FROM use_items GROUP BY note_id) uiq ON uiq.note_id = u.id
        LEFT JOIN receiving_notes r ON r.id = u.receiving_note_id
        WHERE u.status='active' AND {use_where}
        GROUP BY {month_expr('u.date')}, u.project_no, u.project_name, u.warehouse_code, r.seller_name
        ORDER BY {month_expr('u.date')} DESC, u.project_no
    """)
    use_rows = db.session.execute(use_sql, params).fetchall()

    # 构建综合结果
    key_map = {}
    for r in recv_rows:
        key = (r.ym, r.project_no, r.warehouse_code, r.seller_name)
        key_map[key] = {
            "ym": r.ym, "project_no": r.project_no,
            "project_name": r.project_name, "warehouse_code": r.warehouse_code,
            "supplier": r.seller_name,
            "recv_cnt": r.cnt, "recv_qty": round(float(r.qty or 0), 3), "recv_amt": round(float(r.amt or 0), 2),
            "use_cnt": 0, "use_qty": 0, "use_amt": 0,
        }

    for r in use_rows:
        key = (r.ym, r.project_no, r.warehouse_code, r.seller_name)
        if key in key_map:
            key_map[key]["use_cnt"] += r.cnt
            key_map[key]["use_qty"] += round(float(r.qty or 0), 3)
            key_map[key]["use_amt"] += round(float(r.amt or 0), 2)
        else:
            key_map[key] = {
                "ym": r.ym, "project_no": r.project_no,
                "project_name": r.project_name, "warehouse_code": r.warehouse_code,
                "supplier": r.seller_name,
                "recv_cnt": 0, "recv_qty": 0, "recv_amt": 0,
                "use_cnt": r.cnt, "use_qty": round(float(r.qty or 0), 3), "use_amt": round(float(r.amt or 0), 2),
            }

    items = sorted(key_map.values(), key=lambda x: (x["ym"], x["project_no"]), reverse=True)

    # 补充仓库名称
    wh_name_map = get_warehouse_name_map(db.session)
    for item in items:
        item["warehouse_name"] = wh_name_map.get(item["warehouse_code"], "")

    # 汇总统计
    total_recv_amt = sum(i["recv_amt"] for i in items)
    total_use_amt = sum(i["use_amt"] for i in items)
    total_recv_cnt = sum(i["recv_cnt"] for i in items)
    total_use_cnt = sum(i["use_cnt"] for i in items)
    total_recv_qty = round(sum(i["recv_qty"] for i in items), 3)
    total_use_qty = round(sum(i["use_qty"] for i in items), 3)

    # 获取筛选选项（从 unified 接口中获取各维度过滤条件合并到单个 filters 键）
    filter_opts = _get_filter_options(db.session)
    all_years = filter_opts["years"]
    projects = filter_opts["projects"]
    warehouses = filter_opts["warehouses"]
    suppliers = filter_opts["suppliers"]

    # 分组汇总（供前端切换展示）
    # 按仓库
    wh_recv = db.session.execute(text(f"""
        SELECT r.warehouse_code, COUNT(*) as cnt, SUM(r.total_amount) as amt
        FROM receiving_notes r WHERE r.status='active' AND {recv_where}
        GROUP BY r.warehouse_code
    """), params).fetchall()
    wh_use = db.session.execute(text(f"""
        SELECT u.warehouse_code, COUNT(*) as cnt, SUM(u.total_amount) as amt
        FROM use_notes u WHERE u.status='active' AND {use_where}
        GROUP BY u.warehouse_code
    """), params).fetchall()
    wh_map = {}
    for r in wh_recv:
        wh_map[r.warehouse_code or ''] = {"recv_cnt": r.cnt, "recv_amt": round(float(r.amt or 0), 2), "use_cnt": 0, "use_amt": 0}
    for r in wh_use:
        k = r.warehouse_code or ''
        if k not in wh_map:
            wh_map[k] = {"recv_cnt": 0, "recv_amt": 0, "use_cnt": 0, "use_amt": 0}
        wh_map[k]["use_cnt"] += r.cnt
        wh_map[k]["use_amt"] += round(float(r.amt or 0), 2)
    grouped_by_warehouse = [{"code": k, "name": wh_name_map.get(k, k),
        "recv_cnt": v["recv_cnt"], "recv_amt": v["recv_amt"],
        "use_cnt": v["use_cnt"], "use_amt": v["use_amt"]}
        for k, v in sorted(wh_map.items())]

    # 按工程
    proj_recv = db.session.execute(text(f"""
        SELECT r.project_no, r.project_name, COUNT(*) as cnt, SUM(r.total_amount) as amt
        FROM receiving_notes r WHERE r.status='active' AND {recv_where}
        GROUP BY r.project_no, r.project_name
    """), params).fetchall()
    proj_use = db.session.execute(text(f"""
        SELECT u.project_no, u.project_name, COUNT(*) as cnt, SUM(u.total_amount) as amt
        FROM use_notes u WHERE u.status='active' AND {use_where}
        GROUP BY u.project_no, u.project_name
    """), params).fetchall()
    proj_map = {}
    for r in proj_recv:
        proj_map[r.project_no or ''] = {"project_no": r.project_no or '', "project_name": r.project_name or '', "recv_cnt": r.cnt, "recv_amt": round(float(r.amt or 0), 2), "use_cnt": 0, "use_amt": 0}
    for r in proj_use:
        k = r.project_no or ''
        if k not in proj_map:
            proj_map[k] = {"project_no": k, "project_name": r.project_name or '', "recv_cnt": 0, "recv_amt": 0, "use_cnt": 0, "use_amt": 0}
        proj_map[k]["use_cnt"] += r.cnt
        proj_map[k]["use_amt"] += round(float(r.amt or 0), 2)
    grouped_by_project = sorted(proj_map.values(), key=lambda x: -x["recv_amt"])

    # 按供应商
    sup_recv = db.session.execute(text(f"""
        SELECT r.seller_name, COUNT(*) as cnt, SUM(r.total_amount) as amt
        FROM receiving_notes r WHERE r.status='active' AND {recv_where}
        GROUP BY r.seller_name
    """), params).fetchall()
    sup_use = db.session.execute(text(f"""
        SELECT r2.seller_name, COUNT(*) as cnt, SUM(u.total_amount) as amt
        FROM use_notes u LEFT JOIN receiving_notes r2 ON r2.id = u.receiving_note_id
        WHERE u.status='active' AND {use_where}
        GROUP BY r2.seller_name
    """), params).fetchall()
    sup_map = {}
    for r in sup_recv:
        sup_map[r.seller_name or ''] = {"supplier": r.seller_name or '', "recv_cnt": r.cnt, "recv_amt": round(float(r.amt or 0), 2), "use_cnt": 0, "use_amt": 0}
    for r in sup_use:
        k = r.seller_name or ''
        if k not in sup_map:
            sup_map[k] = {"supplier": k, "recv_cnt": 0, "recv_amt": 0, "use_cnt": 0, "use_amt": 0}
        sup_map[k]["use_cnt"] += r.cnt
        sup_map[k]["use_amt"] += round(float(r.amt or 0), 2)
    grouped_by_supplier = sorted([v for v in sup_map.values() if v["supplier"]], key=lambda x: -x["recv_amt"])

    return jsonify({
        "summary": {
            "recv_amt": total_recv_amt,
            "use_amt": total_use_amt,
            "recv_cnt": total_recv_cnt,
            "use_cnt": total_use_cnt,
            "recv_qty": total_recv_qty,
            "use_qty": total_use_qty,
        },
        "items": items,
        "grouped_by_warehouse": grouped_by_warehouse,
        "grouped_by_project": grouped_by_project,
        "grouped_by_supplier": grouped_by_supplier,
        "filters": {
            "years": all_years,
            "projects": projects,
            "warehouses": warehouses,
            "suppliers": suppliers,
        }
    })


@bp.route("/reports/query/export", methods=["GET"])
def report_query_export():
    """导出统一报表 — 与 query 相同筛选条件，输出 Excel"""
    import os
    from flask import current_app
    from services.excel_generator import new_report_wb

    view = request.args.get("view", "detail")

    # 复用查询逻辑获取数据
    resp = report_query()
    data = resp[0] if isinstance(resp, tuple) else resp.get_json()

    items = data.get("items", [])
    summary = data.get("summary", {})

    if view == "warehouse":
        # 收发存汇总表格式（与打印一致）
        now = datetime.now()
        y = request.args.get("years", str(now.year)).split(",")[0]
        m = request.args.get("months", f"{now.month:02d}").split(",")[0]

        bal = _get_warehouse_balance(db.session, y, m)
        recv_before = bal["recv_before"]
        use_before = bal["use_before"]
        recv_month = bal["recv_month"]
        use_month = bal["use_month"]
        ms = bal["month_start"]
        me = bal["month_end"]

        wh_name_map = get_warehouse_name_map(db.session)
        all_codes = sorted(set(list(recv_before) + list(use_before) + list(recv_month) + list(use_month)))

        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side
        wb = Workbook()
        ws = wb.active
        ws.title = "收发存汇总表"
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 18

        tf = Font(name="微软雅黑", size=14, bold=True)
        nf = Font(name="微软雅黑", size=10)
        bf = Font(name="微软雅黑", size=10, bold=True)
        ca = Alignment(horizontal="center", vertical="center")
        ra = Alignment(horizontal="right", vertical="center")
        la = Alignment(horizontal="left", vertical="center")
        tb = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))

        # Row 1: 标题
        ws.merge_cells("A1:E1")
        ws["A1"] = "收发存汇总表"
        ws["A1"].font = tf
        ws["A1"].alignment = ca
        ws.row_dimensions[1].height = 30

        # Row 2: 记账日期
        ws.merge_cells("A2:E2")
        ws["A2"] = f"记账日期：{ms}~{me}"
        ws["A2"].font = nf
        ws["A2"].alignment = la

        # Row 3: 仓库列表（含期初有余额的仓库）
        wh_list_str = "、".join([code + wh_name_map.get(code, code) for code in all_codes])
        ws.merge_cells("A3:E3")
        ws["A3"] = f"仓库：{wh_list_str}"
        ws["A3"].font = nf
        ws["A3"].alignment = la

        # Row 4: 表头
        headers = ["仓库", "期初金额", "收入金额", "发出金额", "结存金额"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=c, value=h)
            cell.font = bf
            cell.alignment = ca
            cell.border = tb

        # 数据行
        total_opening = total_income = total_expense = total_balance = 0
        row_idx = 5
        for code in all_codes:
            opening = recv_before.get(code, 0) - use_before.get(code, 0)
            income = recv_month.get(code, 0)
            expense = use_month.get(code, 0)
            balance = round(opening + income - expense, 2)
            if not opening and not income and not expense and not balance:
                continue
            name = wh_name_map.get(code, code)
            ws.cell(row=row_idx, column=1, value=code + name).font = nf
            ws.cell(row=row_idx, column=1).alignment = ca
            ws.cell(row=row_idx, column=1).border = tb
            for col_idx, val in [(2, opening), (3, income), (4, expense), (5, balance)]:
                c = ws.cell(row=row_idx, column=col_idx, value=val)
                c.font = nf
                c.number_format = '#,##0.00'
                c.alignment = ra
                c.border = tb
            total_opening += opening
            total_income += income
            total_expense += expense
            total_balance += balance
            row_idx += 1

        # 合计行
        ws.cell(row=row_idx, column=1, value="合计").font = bf
        ws.cell(row=row_idx, column=1).alignment = ca
        ws.cell(row=row_idx, column=1).border = tb
        for col_idx, val in [(2, total_opening), (3, total_income), (4, total_expense), (5, total_balance)]:
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = bf
            c.number_format = '#,##0.00'
            c.alignment = ra
            c.border = tb

        # 底部空行 + 单位/制表 + 打印日期
        row_idx += 2
        ws.cell(row=row_idx, column=1, value="单位：天津港航安装工程有限公司").font = nf
        ws.cell(row=row_idx, column=1).alignment = la
        user_name = session.get("display_name") or session.get("username", "")
        ws.cell(row=row_idx, column=4, value=f"制表：{user_name}").font = nf
        ws.cell(row=row_idx, column=4).alignment = ra
        ws.merge_cells(f"A{row_idx}:C{row_idx}")
        ws.merge_cells(f"D{row_idx}:E{row_idx}")
        row_idx += 1
        ws.cell(row=row_idx, column=5, value=f"打印日期：{me}").font = nf
        ws.cell(row=row_idx, column=5).alignment = ra
    elif view == "project":
        headers = ["工程编号", "工程名称", "点收单数", "点收金额", "领用单数", "领用金额"]
        col_widths = [(1, 12), (2, 25), (3, 10), (4, 16), (5, 10), (6, 16)]
        wb, ws, s = new_report_wb("物资点收领用汇总报表（按工程）", headers, col_widths)
        ws.title = "按工程汇总"
        group_data = data.get("grouped_by_project", [])
        total_recv_cnt = total_recv_amt = total_use_cnt = total_use_amt = 0
        for i, item in enumerate(group_data, 3):
            ws.cell(row=i, column=1, value=item.get("project_no", "")).font = s["normal_font"]
            ws.cell(row=i, column=2, value=item.get("project_name", "")).font = s["normal_font"]
            ws.cell(row=i, column=3, value=item.get("recv_cnt", 0)).font = s["normal_font"]
            c4 = ws.cell(row=i, column=4, value=item.get("recv_amt", 0)); c4.font = s["normal_font"]; c4.number_format = '#,##0.00'
            ws.cell(row=i, column=5, value=item.get("use_cnt", 0)).font = s["normal_font"]
            c6 = ws.cell(row=i, column=6, value=item.get("use_amt", 0)); c6.font = s["normal_font"]; c6.number_format = '#,##0.00'
            for col in range(1, 7):
                ws.cell(row=i, column=col).border = s["thin_border"]
            total_recv_cnt += item.get("recv_cnt", 0)
            total_recv_amt += item.get("recv_amt", 0)
            total_use_cnt += item.get("use_cnt", 0)
            total_use_amt += item.get("use_amt", 0)
        row = len(group_data) + 3
        ws.cell(row=row, column=1, value="合计").font = s["bold_font"]
        ws.cell(row=row, column=3, value=total_recv_cnt).font = s["bold_font"]
        c4 = ws.cell(row=row, column=4, value=total_recv_amt); c4.font = s["bold_font"]; c4.number_format = '#,##0.00'
        ws.cell(row=row, column=5, value=total_use_cnt).font = s["bold_font"]
        c6 = ws.cell(row=row, column=6, value=total_use_amt); c6.font = s["bold_font"]; c6.number_format = '#,##0.00'
    elif view == "supplier":
        headers = ["供应商", "点收单数", "点收金额", "领用单数", "领用金额"]
        col_widths = [(1, 25), (2, 10), (3, 16), (4, 10), (5, 16)]
        wb, ws, s = new_report_wb("物资点收领用汇总报表（按供应商）", headers, col_widths)
        ws.title = "按供应商汇总"
        group_data = data.get("grouped_by_supplier", [])
        total_recv_cnt = total_recv_amt = total_use_cnt = total_use_amt = 0
        for i, item in enumerate(group_data, 3):
            ws.cell(row=i, column=1, value=item.get("supplier", "")).font = s["normal_font"]
            ws.cell(row=i, column=2, value=item.get("recv_cnt", 0)).font = s["normal_font"]
            c3 = ws.cell(row=i, column=3, value=item.get("recv_amt", 0)); c3.font = s["normal_font"]; c3.number_format = '#,##0.00'
            ws.cell(row=i, column=4, value=item.get("use_cnt", 0)).font = s["normal_font"]
            c5 = ws.cell(row=i, column=5, value=item.get("use_amt", 0)); c5.font = s["normal_font"]; c5.number_format = '#,##0.00'
            for col in range(1, 6):
                ws.cell(row=i, column=col).border = s["thin_border"]
            total_recv_cnt += item.get("recv_cnt", 0)
            total_recv_amt += item.get("recv_amt", 0)
            total_use_cnt += item.get("use_cnt", 0)
            total_use_amt += item.get("use_amt", 0)
        row = len(group_data) + 3
        ws.cell(row=row, column=1, value="合计").font = s["bold_font"]
        ws.cell(row=row, column=2, value=total_recv_cnt).font = s["bold_font"]
        c3 = ws.cell(row=row, column=3, value=total_recv_amt); c3.font = s["bold_font"]; c3.number_format = '#,##0.00'
        ws.cell(row=row, column=4, value=total_use_cnt).font = s["bold_font"]
        c5 = ws.cell(row=row, column=5, value=total_use_amt); c5.font = s["bold_font"]; c5.number_format = '#,##0.00'
    else:
        # 明细视图（含点收/领用数量合计）
        wb, ws, s = new_report_wb(
            "物资点收领用汇总报表",
            ["月份", "工程编号", "工程名称", "仓库", "供应商",
             "点收单数", "点收数量", "点收金额", "领用单数", "领用数量", "领用金额"],
            [(1, 10), (2, 12), (3, 25), (4, 10), (5, 25),
             (6, 10), (7, 12), (8, 16), (9, 10), (10, 12), (11, 16)]
        )
        ws.title = "汇总报表"
        for i, item in enumerate(items, 3):
            ws.cell(row=i, column=1, value=item.get("ym", ""))
            ws.cell(row=i, column=2, value=item.get("project_no", ""))
            ws.cell(row=i, column=3, value=item.get("project_name", ""))
            ws.cell(row=i, column=4, value=item.get("warehouse_code", "") + item.get("warehouse_name", ""))
            ws.cell(row=i, column=5, value=item.get("supplier", ""))
            ws.cell(row=i, column=6, value=item.get("recv_cnt", 0))
            c7 = ws.cell(row=i, column=7, value=item.get("recv_qty", 0))
            c7.number_format = '#,##0.000'
            c8 = ws.cell(row=i, column=8, value=item.get("recv_amt", 0))
            c8.number_format = '#,##0.00'
            ws.cell(row=i, column=9, value=item.get("use_cnt", 0))
            c10 = ws.cell(row=i, column=10, value=item.get("use_qty", 0))
            c10.number_format = '#,##0.000'
            c11 = ws.cell(row=i, column=11, value=item.get("use_amt", 0))
            c11.number_format = '#,##0.00'
            for col in range(1, 12):
                ws.cell(row=i, column=col).font = s["normal_font"]
                ws.cell(row=i, column=col).border = s["thin_border"]
                if col in (1, 6, 9):
                    ws.cell(row=i, column=col).alignment = s["center_align"]
        row = len(items) + 3
        ws.cell(row=row, column=1, value="合计").font = s["bold_font"]
        ws.cell(row=row, column=6, value=summary.get("recv_cnt", 0)).font = s["bold_font"]
        c7 = ws.cell(row=row, column=7, value=summary.get("recv_qty", 0))
        c7.font = s["bold_font"]; c7.number_format = '#,##0.000'
        c8 = ws.cell(row=row, column=8, value=summary.get("recv_amt", 0))
        c8.font = s["bold_font"]; c8.number_format = '#,##0.00'
        ws.cell(row=row, column=9, value=summary.get("use_cnt", 0)).font = s["bold_font"]
        c10 = ws.cell(row=row, column=10, value=summary.get("use_qty", 0))
        c10.font = s["bold_font"]; c10.number_format = '#,##0.000'
        c11 = ws.cell(row=row, column=11, value=summary.get("use_amt", 0))
        c11.font = s["bold_font"]; c11.number_format = '#,##0.00'

    export_dir = os.path.join(current_app.config.get("UPLOAD_FOLDER", "uploads"), "exports")
    os.makedirs(export_dir, exist_ok=True)
    suffix = {"warehouse": "by_warehouse", "project": "by_project", "supplier": "by_supplier"}.get(view, "detail")
    fname = f"report_{suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = os.path.join(export_dir, fname)
    wb.save(output_path)

    from flask import send_file
    return send_file(output_path, as_attachment=True, download_name=fname)


@bp.route("/reports/warehouse-print", methods=["GET"])
def warehouse_print():
    """收发存汇总表打印数据（按仓库的期初/收入/发出/结存）"""
    now = datetime.now()
    year = request.args.get("year", str(now.year))
    month = request.args.get("month", f"{now.month:02d}")

    # 收发存数据
    bal = _get_warehouse_balance(db.session, year, month)
    recv_before = bal["recv_before"]
    use_before = bal["use_before"]
    recv_month = bal["recv_month"]
    use_month = bal["use_month"]
    month_start = bal["month_start"]
    month_end = bal["month_end"]

    wh_name_map = get_warehouse_name_map(db.session)
    all_codes = sorted(set(list(recv_before.keys()) + list(use_before.keys()) +
                           list(recv_month.keys()) + list(use_month.keys())))

    warehouses = []
    for code in all_codes:
        opening = recv_before.get(code, 0) - use_before.get(code, 0)
        income = recv_month.get(code, 0)
        expense = use_month.get(code, 0)
        balance = round(opening + income - expense, 2)
        warehouses.append({
            "code": code,
            "name": wh_name_map.get(code, code),
            "opening": opening,
            "income": income,
            "expense": expense,
            "balance": balance,
        })

    # 合计行
    total_opening = round(sum(w["opening"] for w in warehouses), 2)
    total_income = round(sum(w["income"] for w in warehouses), 2)
    total_expense = round(sum(w["expense"] for w in warehouses), 2)
    total_balance = round(sum(w["balance"] for w in warehouses), 2)

    # 打印日期固定为筛选当月最后一天
    print_date = month_end

    return jsonify({
        "month_start": month_start,
        "month_end": month_end,
        "print_date": print_date,
        "warehouses": warehouses,
        "total": {"opening": total_opening, "income": total_income, "expense": total_expense, "balance": total_balance},
    })


# ===== 操作日志 =====

@bp.route("/logs/export", methods=["GET"])
def export_logs():
    """导出操作日志到 Excel（仅管理员）"""
    if session.get("role") != "admin":
        return jsonify({"error": "仅管理员可导出操作日志"}), 403
    import openpyxl
    from flask import send_file
    from io import BytesIO

    logs = OperationLog.query.order_by(OperationLog.id.desc()).all()
    ACTION_LABEL = {"create": "创建", "update": "修改", "delete": "删除", "export": "导出"}
    TARGET_LABEL = {"receiving_note": "点收单", "use_note": "领用单", "invoice": "发票",
                    "project": "工程", "personnel": "人员", "supplier": "供应商",
                    "warehouse": "仓库", "user": "用户"}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "操作日志"
    ws.append(["ID", "用户", "操作", "对象类型", "对象ID", "详情", "时间"])
    for l in logs:
        ws.append([
            l.id, l.username,
            ACTION_LABEL.get(l.action, l.action),
            TARGET_LABEL.get(l.target_type, l.target_type),
            l.target_id or "",
            l.detail or "",
            l.created_at or "",
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="操作日志.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.route("/logs", methods=["GET"])
def list_logs():
    """获取操作日志（仅管理员）"""
    if session.get("role") != "admin":
        return jsonify({"error": "仅管理员可查看操作日志"}), 403
    page = request.args.get("page", 1, type=int)
    limit = request.args.get("limit", 50, type=int)
    keyword = request.args.get("keyword", "").strip()

    query = OperationLog.query.order_by(OperationLog.id.desc())
    if keyword:
        like = f"%{keyword}%"
        query = query.filter(
            db.or_(OperationLog.username.like(like),
                   OperationLog.action.like(like),
                   OperationLog.target_type.like(like),
                   OperationLog.detail.like(like))
        )

    total = query.count()
    logs = query.offset((page - 1) * limit).limit(limit).all()

    return jsonify({
        "items": [{
            "id": l.id,
            "user_id": l.user_id,
            "username": l.username,
            "action": l.action,
            "action_label": {"create": "创建", "update": "修改", "delete": "删除", "export": "导出"}.get(l.action, l.action),
            "target_type": l.target_type,
            "target_type_label": {"receiving_note": "点收单", "use_note": "领用单", "invoice": "发票", "project": "工程", "personnel": "人员", "supplier": "供应商", "warehouse": "仓库", "user": "用户"}.get(l.target_type, l.target_type),
            "target_id": l.target_id,
            "detail": l.detail,
            "created_at": l.created_at,
        } for l in logs],
        "total": total,
        "page": page,
        "limit": limit,
    })


# ===== 系统参数配置 =====

@bp.route("/settings", methods=["GET"])
def list_settings():
    settings = SystemConfig.query.all()
    return jsonify({s.key: s.value for s in settings})


@bp.route("/settings", methods=["PUT"])
def update_settings():
    """更新系统参数（仅 admin）"""
    from flask import session
    if session.get("role") != "admin":
        return jsonify({"error": "仅管理员可执行此操作"}), 403
    data = request.get_json()
    for key, value in data.items():
        cfg = SystemConfig.query.filter_by(key=key).first()
        if cfg:
            cfg.value = str(value)
        else:
            db.session.add(SystemConfig(key=key, value=str(value)))
    log_operation("update", "setting", 0, "更新系统参数")
    db.session.commit()
    return jsonify({"message": "参数已更新"})


# ===== 文件清理 =====

@bp.route("/cleanup/scan", methods=["GET"])
def scan_orphans():
    """扫描孤儿文件（仅 admin）"""
    from flask import session
    if session.get("role") != "admin":
        return jsonify({"error": "仅管理员可执行此操作"}), 403
    from services.file_cleaner import scan_orphan_files
    result = scan_orphan_files(os.path.join(os.path.dirname(os.path.dirname(__file__)), "uploads"))
    return jsonify({
        "orphan_count": len(result["orphans"]),
        "total_size_mb": round(result["total_size"] / 1024 / 1024, 2),
        "orphans": [os.path.basename(f) for f in result["orphans"]],
    })


@bp.route("/cleanup/execute", methods=["POST"])
def execute_cleanup():
    """执行孤儿文件清理（仅 admin）"""
    from flask import session
    if session.get("role") != "admin":
        return jsonify({"error": "仅管理员可执行此操作"}), 403
    data = request.get_json() or {}
    max_age_days = data.get("max_age_days", 90)
    from services.file_cleaner import clean_orphan_files
    result = clean_orphan_files(
        os.path.join(os.path.dirname(os.path.dirname(__file__)), "uploads"),
        max_age_days=max_age_days
    )
    log_operation("delete", "file", 0,
                  f"清理孤儿文件 {result['deleted']} 个，释放 {result['freed_mb']}MB")
    db.session.commit()
    return jsonify(result)


# ===== 基础数据初始化 =====

@bp.route("/init-data", methods=["GET"])
def get_init_data():
    """前端需要的基础数据（仓库列表、人员等）"""
    from services.warehouse_matcher import get_warehouse_list
    warehouses = get_warehouse_list()
    projects = [{"no": p.project_no, "name": p.project_name}
                for p in sorted(Project.query.all(), key=_project_sort_key)]
    accountants = [{"name": p.name} for p in Personnel.query.filter(Personnel.role.contains("accountant"), Personnel.enabled.is_(True)).all()]
    buyers = [{"name": p.name} for p in Personnel.query.filter(Personnel.role.contains("buyer"), Personnel.enabled.is_(True)).all()]
    recipients = [{"name": p.name} for p in Personnel.query.filter(Personnel.role.contains("recipient"), Personnel.enabled.is_(True)).all()]

    return jsonify({
        "warehouses": warehouses,
        "projects": projects,
        "accountants": accountants,
        "buyers": buyers,
        "recipients": recipients,
    })
